import logging
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import numpy as np
import os
import io
import re
from datetime import datetime, timedelta
import azure.functions as func
from azure.storage.blob import BlobServiceClient
from openpyxl.styles import PatternFill, Font, Alignment
import tempfile

def get_blob_df(container_name, file_name, sheet_name=None)
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
    blob_data = io.BytesIO(blob_client.download_blob().content_as_bytes())
    try
        if sheet_name
            current_df_pd = pd.read_excel(blob_data, sheet_name=sheet_name, index_col=None)
        else
            current_df_pd = pd.read_excel(blob_data, index_col=None)
    except
        current_df_pd = pd.read_csv(blob_data, index_col=None)
    return current_df_pd

def upload_file(file_content, container_name, file_name)
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
    blob_client.upload_blob(file_content, overwrite=True)
    return

def format_dates(df, columns, date_format)
    for column in columns
        if column in df.columns
            df[column] = pd.to_datetime(df[column], errors='coerce').dt.strftime(date_format)
    return df

def excel_formatting(path)
    df_dict = pd.read_excel(path, sheet_name=None)
    writer = pd.ExcelWriter(path, engine='xlsxwriter',
                            date_format='mmddyyyy',
                            datetime_format='mmddyyyy')

    for key, value in df_dict.items()
        name = key
        df = value
        df.to_excel(writer, sheet_name=name, index=False)
    writer.close()
    wb = load_workbook(path)

    for sheet in wb.worksheets
        for col in sheet.iter_cols(min_col=None, max_row=1)
            for i in col
                i.font = Font(color='FFFFFF')
                i.fill = PatternFill(fgColor='000080', fill_type='solid')

        for column_cells in sheet.columns
            sheet.column_dimensions[openpyxl.utils.get_column_letter(
                column_cells[0].column)].auto_size = True

    wb.save(path)
    wb.close()
    return True

def create_excel_file(df)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp
        temp_path = temp.name
        logging.info(f'Creating Excel file at {temp_path}')
        writer = pd.ExcelWriter(
            temp_path, engine='xlsxwriter',
            date_format='mmddyyyy',
            datetime_format='mmddyyyy'
        )
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

    excel_formatting(temp_path)
    return temp_path

def add_suffix_columns(df, suffix=suffix)
    columns = df.columns.tolist()
    new_columns = []
    for column in columns
        new_columns.append(column)
        if last_name in column and spouse_last_name not in column and other not in column
            new_column_name = suffix
            new_columns.append(new_column_name)
            df[new_column_name] = 
        elif spouse_last_name in column
            new_column_name = spouse_suffix
            new_columns.append(new_column_name)
            df[new_column_name] = 
        for i in range(1, 11)
            if fother_{i}_last_name in column
                new_column_name = fother_{i}_{suffix}
                new_columns.append(new_column_name)
                df[new_column_name] = 
                break 
    df = df[new_columns]  
    return df

def main(req func.HttpRequest) - func.HttpResponse
    logging.info('Python HTTP trigger function processed a request.')

    global connection_string
    connection_string = os.environ[CONNECTION_STRING]

    current_datetime = datetime.now()
    one_day = timedelta(days=1)
    previous_datetime = current_datetime - one_day
    formatted_date = previous_datetime.strftime('%#m-%#d-%y')

    req_body = req.get_json()
    file_name = req_body.get('name')
    logging.info(f'{file_name}')
    
    df = get_blob_df('sftp-content', f'ReportsEnrollment ReportSponsored Tab{file_name}')

    if 'agent__c' in df.columns
        df['agent__c'] = 'Neal, Brian' 
        df['npn_used__c'] = '18152472'
        df['npn_reason__c'] = df['npn_reason__c'].replace('agent_override', 'agent')
        df['last_submission_date__c'] = pd.to_datetime(df['last_submission_date__c'], errors='coerce')

    date_columns = [
            'dob', 'last_submission_date', 'effective_date', 'spouse_dob', 'other_1_dob', 
            'other_2_dob', 'other_3_dob', 'other_4_dob', 'other_5_dob', 
            'other_6_dob', 'other_7_dob', 'other_8_dob', 'other_9_dob', 
            'other_10_dob'
        ]
    rename_columns_df = {
            'Enrollment_Report_Disposition' 'Disposition',
            'SF_Primary_Contact__r.Social_Security''ssn',
            'SF_Spouse_Contact__r.Social_Security''spouse_ssn',
            'SF_Dependent_1_Contact__r.Social_Security''other_1_ssn',
            'SF_Dependent_2_Contact__r.Social_Security''other_2_ssn'
        }
    
    rename_columns = {col col.replace('__c', '') for col in df.columns}
    df = df.rename(columns=rename_columns)
    df = df.rename(columns=rename_columns_df)
    df = format_dates(df, date_columns, '%#m%#d%Y')
    df['last_submission_date'] = pd.to_datetime(df['last_submission_date'], format='%m%d%Y')
    df['spouse_dob'] = pd.to_datetime(df['spouse_dob'], errors='coerce').dt.strftime('%#m%#d%Y')
    df.sort_values(by=['last_submission_date'], inplace=True, ascending=False)
    df['last_submission_date'] = df['last_submission_date'].dt.strftime('%#m%#d%Y')
    df = add_suffix_columns(df)

    if 'ssn' in df.columns
        df['ssn'] = df['ssn'].str.replace('-', '', regex=False)

    if 'spouse_first_name' in df.columns and 'spouse_applying' in df.columns
        df.loc[df['spouse_first_name'].isna()  (df['spouse_first_name'] == ''), 'spouse_applying'] = ''


    for i in range(1, 11)
        first_name_col = f'other_{i}_first_name'
        applying_col = f'other_{i}_applying'
        if first_name_col in df.columns and applying_col in df.columns
            logging.info(f'Processing {applying_col}')
            df.loc[df[first_name_col].isna()  (df[first_name_col] == ''), applying_col] = ''
            df[applying_col] = df[applying_col].replace({0 'FALSE', 1 'TRUE'})


    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp
        temp_path = temp.name
        writer = pd.ExcelWriter(
            temp_path, engine='xlsxwriter',
            date_format='mmddyyyy',
            datetime_format='mmddyyyy'
        )
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()

    excel_formatting(temp_path)

    with open(temp_path, 'rb') as file
        file_content = file.read()

    try
        Enorllment_report = f'HH OEP ENROLLMENTS {formatted_date}.xlsx'
        upload_file(file_content, 'sftp-content', fReportsEnrollment ReportProductionSponsored Tab{Enorllment_report})
    finally
        os.remove(temp_path)
        logging.info(fTemporary file {temp_path} removed)


    return func.HttpResponse(fEnrollment report created {Enorllment_report}, status_code=200)